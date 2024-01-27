import os
import shutil
from oauth2client.service_account import ServiceAccountCredentials
import gspread
from cfg.database import Database
from gspread.utils import ExportFormat
import openpyxl


db = Database('cfg/database')


def overwriting_data():
    link = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
    my_creds = ServiceAccountCredentials.from_json_keyfile_name("cfg/CarDpAPI.json", link)
    client = gspread.authorize(my_creds)

    all_values_sheet = client.open('Car_dp').get_worksheet(0).get_all_values()[1::]
    db.del_all_users()
    db.add_dada_sheet(all_values_sheet)

    shutil.rmtree("all_file_excel")
    os.mkdir("all_file_excel")

    all_sheets = client.open("Car_dp").export(format=ExportFormat.EXCEL)
    with open('all_file_excel/output.xlsx', 'wb') as f:
        f.write(all_sheets)
    
    wb = openpyxl.load_workbook('all_file_excel/output.xlsx')
    sheets = wb.sheetnames
    wb.save('all_file_excel/output.xlsx')

    for idx, s in enumerate(sheets):
        wb = openpyxl.load_workbook('all_file_excel/output.xlsx')
        sheets = wb.sheetnames

        if idx != 0:
            for idxs, sheet in enumerate(sheets):
                if idx != idxs:
                    sheet_name = wb.get_sheet_by_name(sheets[idxs])
                    wb.remove_sheet(sheet_name)
            wb.save(f'all_file_excel/{sheets[idx]} Всё время.xlsx')

    os.remove("all_file_excel/output.xlsx")

    all_file = os.listdir("all_file_excel")
    for file in all_file:
        if file[-14:-1] == "Всё время.xls":
            # Неделя
            wb = openpyxl.load_workbook(f"all_file_excel/{file}")
            sheet = wb.active

            rows_to_delete = list(range(6, sheet.max_row - 1))

            for row in reversed(rows_to_delete):
                if sheet.cell(row, 1).value is not None:
                    if row != sheet.max_row:
                        if row - 1 > 5:
                            sheet.delete_rows(row - 1)

            wb.save(f'all_file_excel/{file[:-15]} Неделя.xlsx')

            # Месяц
            wb = openpyxl.load_workbook(f"all_file_excel/{file}")
            sheet = wb.active

            rows_to_delete = list(range(6, sheet.max_row - 4))

            for row in reversed(rows_to_delete):
                if sheet.cell(row, 1).value is not None:
                    if row != sheet.max_row:
                        if row - 4 > 5:
                            sheet.delete_rows(row - 4)

            wb.save(f'all_file_excel/{file[:-15]} Месяц.xlsx')
